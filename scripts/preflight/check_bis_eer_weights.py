"""
scripts/preflight/check_bis_eer_weights.py

ADD -- closes (partially -- see "What this script cannot resolve" below)
Gate 1 in the ADR registry: "ADR-017: Exact Broad Dollar basket weights --
PARTIALLY GATED -- blocked on BIS EER weight-component data availability
(Gate 1). Not implemented as concrete numbers" and the identical ADR-018
gate for the IDR basket-weight override magnitude.

UPD (alpha-factory_preflight_logs, 28 July 2026): this script was actually
run and found every request 404ing (data) or 501ing (--discover). Fixed
at the time to the /api/v2/ path structure -- necessary, but as of the
29 July live re-run, NOT sufficient (still 404/501). See the FIX BIS-1
block above the endpoint constants for the actual root cause found 1 Aug
2026: the dataflow ID itself was wrong (WS_EER_M -> WS_EER) and the
--discover query was missing the "structure/" path segment entirely
(explaining the 501, which is a different failure signature than the
plain 404 a bad key alone would produce). WS_EER_D still does not appear
anywhere in confirmed evidence (BIS's own portal only shows Monthly EER
across every country sampled) -- WS_EER remains the only target, now
correctly named.

UPD (Ovi, same thread): MXN removed from the currency set below,
replaced with IDR. MXN (Mexican Peso) was carried over unmodified from
Architecture v2.0 §7.2's original BIS_WEIGHTS dict -- a generic "EM
currency" placeholder from before this platform's Indonesia-specific work
(ADR-013 through ADR-018) existed. It has zero relevance to this
platform and was never actually used anywhere else in the codebase
(confirmed: this script was MXN's only occurrence in the entire repo).
IDR is the economically relevant EM currency here -- already a Layer 1
forex pair (USD_IDR), already referenced in instruments_taxonomy.yaml's
own comments as part of the *current* Broad Dollar basket design
("plus USD_IDR (Layer 1, weight logic changes per ADR-018)"), and BI
(Bank Indonesia) is already BIS-covered via context_rates_em_cb.

RESOLVED (Ovi, this thread, following up on the flag below): HKD/TWD/NOK
added to BROAD_DOLLAR_REF_AREAS. The *current* Broad Dollar basket design
(per instruments_taxonomy.yaml's own dollar/dollar_basket comments) is
EUR/JPY/GBP/CAD/CHF/AUD + IDR (all Layer 1) + the 6-currency
context_dollar_basket group (CNH/KRW/SGD/HKD/TWD/NOK) -- 13 currencies.
This script previously covered only 10 -- flagged as a known gap rather
than guessed at ("Ovi's instruction was specifically MXN->IDR") pending
explicit instruction, which has now been given. All 13 currencies are
covered as of this thread; BIS_EER_ENDPOINT_MONTHLY's key is now built
FROM BROAD_DOLLAR_REF_AREAS.values() rather than hand-duplicated, so this
class of drift-between-dict-and-key cannot recur (see the FIX comment
above that constant).

Same authoring/execution split as the other three scripts in this
directory: this sandbox's network allowlist has no route to
stats.bis.org. Authoring does not require network access; running it
does.

UPD (Ovi, 3 Aug 2026 thread): Gate 1 substantially advanced. BIS's own
data.bis.org/topics/EER page (server-rendered, unlike the SPA pages
elsewhere on this project) links directly, under its own "Methodology"
section, to a downloadable weights table:
https://www.bis.org/statistics/eer/weightsb.xlsx (Broad, 64 economies --
the one relevant here, since Narrow only covers 26/27 core economies and
IDR/HKD/TWD would not be in it). Confirmed reachable and genuinely an
.xlsx (mime type application/vnd.openxmlformats-officedocument.
spreadsheetml.sheet, not a redirect or error page) -- Gate 1's weight
COMPONENTS are real and machine-readable, just not via the SDMX API
check_bis_eer_weights.py otherwise uses; they're a plain file download.
Also confirmed via the same page: weights are TIME-VARYING on a 3-year
basis (vintages 1993-95 through 2017-19 per BIS's own FAQ; the 2017-19
vintage has been in continuous use for "the latest period" since, until
BIS publishes the next 3-year update) -- there is no single permanent
"exact weight," but there IS a specific, nameable current vintage, which
is what --discover-weights below is for. See _discover_weights() -- same
two-phase discover-then-extract pattern as --discover for the API
structure, since this thread cannot inspect the file's actual internal
layout (no network route to bis.org from any sandbox on this project;
Ovi's next run on the M1 is what actually reveals row/column structure
for a targeted extraction pass).

UPD (Ovi, M1, 4 Aug 2026): --discover-weights run for real. weightsb.xlsx
downloaded clean (492,941 bytes), 10 sheets (1993_1995 through 2020_2022
-- confirms the 3-year vintage cadence directly; no newer vintage exists
yet). Every sheet is a symmetric "who weights whom" matrix: row label
(column 2) = country's own REF_AREA code ("In the EER for:"), column
header (row 6) = REF_AREA code being weighted ("Weight on:"), cell =
percent weight. All 13 BROAD_DOLLAR_REF_AREAS codes found at IDENTICAL
(row, col) positions in EVERY one of the 10 sheets -- layout is now fully
characterized, not just located. This did NOT yet give the actual weight
VALUES -- the scan finds coordinates for the 13 target currencies as
column headers, but never surfaces the US row itself (US is not one of
the 13 target REF_AREA values, so it was never a search target). See
extract_us_weights_from_sheet() / --extract-weights below.

ADD (this thread): --extract-weights -- the actual Gate 1 answer. Locates
the row where column 2 holds "US" (not found by --discover-weights,
which only searches for the 13 target currency codes), then reads that
row's values at each of the 13 target-currency columns -- i.e. how much
weight each of EUR/JPY/GBP/CAD/CHF/AUD/IDR/CNH/KRW/SGD/HKD/TWD/NOK
carries in the US's own Broad EER basket. This is the empirical
replacement for Architecture v2.0 §7.2's hand-approximated 10-pair
BIS_WEIGHTS dict (which was explicitly labelled "BIS Approximate Trade
Weights" and only covered 10 of the 13 currencies this platform's
current Broad Dollar basket design actually uses).

extract_us_weights_from_sheet() is the pure, testable half (no network,
no file I/O) -- deliberately does NOT hardcode the row 6 / column 2
positions the 4 Aug discovery run found, even though that run reported
them as identical across all 10 sheets: re-deriving per call is the same
"don't guess a layout, verify it" discipline this whole Gate 1 thread has
followed since the WS_CBPOL_D / WS_EER_M / missing "structure/" segment
mistakes, and costs nothing extra (still a single bounded scan of <=200
rows). _extract_weights() is the I/O half: downloads the file, selects a
sheet (defaults to the most recent vintage -- 2020_2022, per the 4 Aug
run -- via max(sheetnames), not the discovery run's confirmed-identical-
across-sheets claim), and prints a clean per-currency report.

Like _discover_weights(), this has been authored and unit-tested against
a SYNTHETIC in-memory workbook only -- no sandbox on this project has a
route to bis.org, so this has not yet been run against the real
weightsb.xlsx file. That is the next step on real hardware.

TYPE decision (Ovi, this thread -- was previously left wildcarded,
pending): NOMINAL, not Real. Two independent reasons converged: (1) DXY
itself -- the index this platform's Broad Dollar Index is explicitly
designed as a companion/extension of (Architecture v2.0 §7.2) -- is a
nominal currency-value index, not inflation-adjusted; using Real EER for
Broad Dollar while DXY stays Nominal would compare two conceptually
different things under one "Dollar strength" umbrella. (2) BIS's own EER
overview page states Daily-frequency EER data exists ONLY for Nominal
indices, never Real ("the latter available only as nominal indices") --
since this platform's Layer 2 anchors are specified at Daily cadence
(Architecture v2.0 §7.2: "Cadence: Daily (same as forex)"), Nominal is
the only choice that can actually deliver that. FREQ is now wildcarded
(empty segment) rather than hardcoded to M or D -- mirroring the exact
same reasoning already applied to check_bis_cbpol_d.py's key: request
whatever frequency BIS actually has per country rather than assume, and
let genuinely-available daily data come through where it exists without
risking a false failure on currencies that may only have monthly EER.
Live-confirmed 4 Aug 2026 (all 13 currencies PASS at 3,813,875 bytes per
currency, ~16x the prior monthly-restricted figure -- see KNOWN_RISKS.md
RISK-16).

Usage:
    python scripts/preflight/check_bis_eer_weights.py
    python scripts/preflight/check_bis_eer_weights.py --discover
    python scripts/preflight/check_bis_eer_weights.py --discover-weights
    python scripts/preflight/check_bis_eer_weights.py --extract-weights
    python scripts/preflight/check_bis_eer_weights.py --extract-weights --sheet 2017_2019

Exit code 0 = EER index data returned for all currencies in the Broad
Dollar basket. Exit code 1 = at least one reference area failed, or
(--discover) no weight-bearing series could be identified in the
dataflow structure, or (--discover-weights / --extract-weights) the
weights file could not be downloaded/parsed, or (--extract-weights) the
US row or a target currency's column could not be located.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))

# FIX (Ovi, this thread -- "issues even though .env already filled"):
# python-dotenv is a declared dependency but was never actually called
# anywhere in this repo (confirmed by grep). Without this, os.getenv()
# only sees variables the shell has separately exported -- a filled
# .env file alone was not enough, which is exactly what this preflight
# run surfaced for TV_USERNAME/TV_PASSWORD and FINNHUB_API_KEY.
from dotenv import load_dotenv
load_dotenv()

# FIX BIS-1 (Ovi, 1 Aug 2026): same root-cause class as check_bis_cbpol_d.py
# -- the 28 July v1->v2 path fix was necessary but not sufficient. Two
# further errors, both fixed here:
#
# 1. Dataflow ID: WS_EER, not WS_EER_M. The "_M" was a monthly-cadence
#    label mistaken for part of the identifier -- same pattern as
#    WS_CBPOL_D, confirmed independently against data.bis.org's own
#    indexed pages ("topics/EER/BIS,WS_EER,1.0/{FREQ}.{TYPE}.{BASKET}.
#    {REF_AREA}", 7 countries checked: US/AE/CN/KR/XM/JP, both Real and
#    Nominal baskets seen). The old WS_EER_M name traces back to a v1-era
#    academic example (fgeerolf.com) whose flow name was itself already a
#    guess from before this project's v1->v2 migration -- both the flow
#    name and the path needed correcting together, not just the path.
# 2. Structure/discovery queries use a DIFFERENT prefix than data queries
#    -- "/api/v2/structure/dataflow/..." not "/api/v2/dataflow/...". The
#    old URL was missing "structure/" entirely, which is consistent with
#    the 501 (not 404) it was actually returning -- a malformed/
#    unrecognized v2 path, not a clean "not found". Confirmed via a real
#    SDMX 2025 conference paper (sdmx2025.org) showing both the data
#    query and structure query shapes for a third sibling dataflow
#    (WS_XRU), independently agreeing with the WS_CBTA data-query example
#    above.
#
# Key = {FREQ}.{TYPE}.{BASKET}.{REF_AREA}. FREQ=M (monthly is what BIS
# publishes EER at -- confirmed across all 7 sampled countries, no daily
# EER variant found). TYPE left wildcarded (empty -- returns both Real
# and Nominal) since neither this platform's Broad Dollar Index design
# nor Gate 1 has settled which one it wants; narrowing that is a separate
# call, not bundled into this endpoint fix. BASKET=B (broad), matching
# "Broad Dollar Index" directly. CONFIRMED LIVE (Ovi, M1, this thread):
# --discover fetched 568951 bytes of real dataflow structure; the data
# query returned 182410 bytes with all 10 (now 13) REF_AREA codes
# present.
#
# FIX (Ovi, this thread): endpoint key is now BUILT FROM
# BROAD_DOLLAR_REF_AREAS.values() rather than a separately hardcoded
# literal string. The literal-string version this replaces is exactly
# how the HKD/TWD/NOK gap happened in the first place -- and how it could
# have silently recurred: adding entries to the dict alone, without this
# change, would leave them permanently unfetched while _check_one() keeps
# confidently reporting "not present" -- indistinguishable from a genuine
# API failure. Building the key from the dict's own values makes this
# whole bug class structurally impossible from here on; the dict is now
# the single source of truth.
# Architecture v2.0 §7.2 BIS_WEIGHTS currencies, MXN removed / IDR added
# (28 Jul 2026); HKD/TWD/NOK added (Ovi, this thread) to complete the
# *current* Broad Dollar basket design per instruments_taxonomy.yaml's
# dollar/dollar_basket comments -- previously flagged as a known gap
# ("Ovi's instruction was specifically MXN->IDR") rather than guessed at,
# now closed on explicit instruction. This dict is the single source of
# truth for BIS_EER_ENDPOINT_MONTHLY's key below (built from its values,
# not hand-duplicated) and for every currency _check_one() validates.
# Mapped to BIS REF_AREA codes matching config/bis_cb_rates.yaml's
# ref_area_map convention.
BROAD_DOLLAR_REF_AREAS: dict[str, str] = {
    "EUR": "XM",  # Euro area
    "JPY": "JP",
    "GBP": "GB",
    "CAD": "CA",
    "CHF": "CH",
    "AUD": "AU",
    "IDR": "ID",  # UPD: replaces MXN -- Indonesia, already covered via context_rates_em_cb (BI)
    "CNH": "CN",  # onshore CNY EER is BIS's only option; CNH itself isn't a BIS REF_AREA
    "KRW": "KR",
    "SGD": "SG",
    "HKD": "HK",  # NEW (Ovi, this thread) -- Hong Kong, pegged currency (reliability_flag pattern, same as SSEC/BOJ YCC)
    "TWD": "TW",  # NEW -- Taiwan
    "NOK": "NO",  # NEW -- Norway
}

# FIX (Ovi, this thread): endpoint key is built FROM BROAD_DOLLAR_REF_AREAS
# .values() -- not a separately hardcoded literal. A hardcoded literal is
# exactly how the HKD/TWD/NOK gap happened in the first place, and how it
# could silently recur: adding entries to the dict alone, without this,
# leaves them permanently unfetched while _check_one() keeps confidently
# reporting "not present" -- indistinguishable from a genuine API
# failure. This makes that whole bug class structurally impossible from
# here on; the dict above is the only place currency membership is
# declared. CONFIRMED LIVE (Ovi, M1, 1 Aug 2026, 10-currency M..B. key):
# --discover fetched 568951 bytes of real dataflow structure; the data
# query returned 182410 bytes with all 10 REF_AREA codes present.
# RE-CONFIRMED LIVE (Ovi, M1, 3 Aug 2026, full 13-currency version): all
# 13 REF_AREA codes PASS, 237188 bytes returned per check -- the HKD/TWD/
# NOK expansion is now empirically confirmed working live, not just
# code-fixed and test-verified.
#
# UPD (Ovi, same 3 Aug thread): renamed from BIS_EER_ENDPOINT_MONTHLY --
# TYPE is now fixed to N (Nominal, see module docstring for the two-part
# rationale) and FREQ is now wildcarded rather than hardcoded to M, so
# "_MONTHLY" was no longer an accurate name. Key changed from "M..B." to
# ".N.B." -- FREQ wildcarded (was fixed M), TYPE fixed to N (was
# wildcarded). LIVE-RECONFIRMED 4 Aug 2026 against this exact key shape
# (3,813,875 bytes/currency, ~16x the prior monthly-only figure).
BIS_EER_ENDPOINT = (
    "https://stats.bis.org/api/v2/data/dataflow/BIS/WS_EER/1.0/"
    ".N.B." + "+".join(BROAD_DOLLAR_REF_AREAS.values())
)
BIS_EER_DATAFLOW_STRUCTURE_URL = "https://stats.bis.org/api/v2/structure/dataflow/BIS/WS_EER/1.0"

# Gate 1 (ADR-017/018): the actual weight COMPONENTS, found this thread --
# not a queryable SDMX series, but a plain file download linked from BIS's
# own methodology page (data.bis.org/topics/EER). Broad (64 economies),
# not Narrow (26/27 -- would not include IDR/HKD/TWD). Confirmed reachable
# and genuinely an .xlsx this thread (mime type verified); internal layout
# now fully characterized (Ovi, M1, 4 Aug 2026 --discover-weights run --
# see module docstring) -- extract_us_weights_from_sheet() / --extract-
# weights below turn that layout into actual per-currency values.
BIS_EER_WEIGHTS_BROAD_URL = "https://www.bis.org/statistics/eer/weightsb.xlsx"

# extract_us_weights_from_sheet() / _extract_weights() (Gate 1 targeted
# extraction, ADD this thread): a row is treated as the "In the EER for:"
# row-label match when its column-2 cell equals this code. A row is
# treated as the "Weight on:" header row when it contains at least this
# many of the 13 target REF_AREA codes as cell values -- 2 rather than 1
# to reliably distinguish the header row (13 hits in the real file, per
# the 4 Aug discovery run) from an ordinary data row, whose column-2 own-
# country code can coincidentally match at most 1 target code and whose
# remaining cells are floats, never REF_AREA strings.
US_REF_AREA = "US"
_HEADER_ROW_MIN_HITS = 2
_EXTRACT_WEIGHTS_MAX_SCAN_ROWS = 200


def extract_us_weights_from_sheet(
    ws,
    ref_areas: dict[str, str] = None,
    us_ref_area: str = US_REF_AREA,
    max_scan_rows: int = _EXTRACT_WEIGHTS_MAX_SCAN_ROWS,
) -> dict[str, float | None] | None:
    """Pure extraction logic for one openpyxl worksheet -- no network, no
    file I/O -- so this can be unit-tested against a synthetic workbook
    (same separation _discover_weights()'s own scan established, applied
    one level further here).

    Locates the row whose column-2 cell equals `us_ref_area` ("In the EER
    for: United States"), separately locates the header row where >= 2 of
    `ref_areas`' values appear as cell values ("Weight on: <currency>"),
    then reads the intersection: the US row's value at each target
    currency's column. Both positions are re-derived from the sheet's own
    content on every call -- never assumed from a fixed row/column number
    -- even though the 4 Aug 2026 discovery run found them identical
    (row 6 header, column 2 labels) across all 10 real vintage sheets.
    Guessing a fixed layout is exactly the failure class this whole Gate 1
    thread exists to avoid (WS_CBPOL_D, WS_EER_M, the missing "structure/"
    segment -- three confident-but-wrong guesses already found and
    corrected elsewhere in this same script's history).

    Returns None if no row's column-2 cell matches `us_ref_area` at all --
    the caller cannot proceed without it. Otherwise returns a dict keyed
    by every key in `ref_areas`, with value None for any currency whose
    column could not be located in the header row or whose US-row cell
    was empty/non-numeric -- partial results are surfaced, not hidden,
    and it is the caller's decision whether a partial result is
    acceptable.

    Caveat found while smoke-testing against a synthetic workbook: the
    header-row heuristic (>= _HEADER_ROW_MIN_HITS matching codes in one
    row) needs `ref_areas` to supply at least 2 currencies to reliably
    tell the true header row apart from an ordinary data row (whose own
    column-2 country code can itself count as a single incidental hit).
    Not a real constraint in practice -- every actual caller passes the
    full 13-currency BROAD_DOLLAR_REF_AREAS -- but relevant if this is
    ever called with a single-currency filter.
    """
    if ref_areas is None:
        ref_areas = BROAD_DOLLAR_REF_AREAS

    scan_rows = min(max_scan_rows, ws.max_row or 0)
    if not scan_rows:
        return None

    rows: dict[int, tuple] = {
        row_idx: row
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1
        )
    }

    us_row_idx = next(
        (
            idx for idx, row in rows.items()
            if len(row) >= 2 and row[1] is not None and str(row[1]).strip() == us_ref_area
        ),
        None,
    )
    if us_row_idx is None:
        return None

    target_ref_areas = set(ref_areas.values())
    col_for_ref_area: dict[str, int] = {}
    for row in rows.values():
        hits = {
            col_idx: str(cell).strip()
            for col_idx, cell in enumerate(row, start=1)
            if cell is not None and str(cell).strip() in target_ref_areas
        }
        if len(hits) >= _HEADER_ROW_MIN_HITS:
            for col_idx, code in hits.items():
                col_for_ref_area.setdefault(code, col_idx)
            break  # first qualifying header row wins -- matches the single stable header row the 4 Aug run found per sheet

    us_row = rows[us_row_idx]
    weights: dict[str, float | None] = {}
    for currency, ref_area in ref_areas.items():
        col_idx = col_for_ref_area.get(ref_area)
        if col_idx is None or col_idx > len(us_row):
            weights[currency] = None
            continue
        cell = us_row[col_idx - 1]  # us_row is a 0-indexed tuple; col_idx is 1-indexed (matches _discover_weights()'s own convention)
        weights[currency] = float(cell) if isinstance(cell, (int, float)) else None

    return weights


def _fetch_csv(url: str) -> str:
    import httpx
    resp = httpx.get(
        url,
        params={"format": "csv", "detail": "dataonly"},
        timeout=30.0,
    )
    resp.raise_for_status()
    return resp.text


def _check_one(ref_area: str) -> tuple[bool, str]:
    try:
        text = _fetch_csv(BIS_EER_ENDPOINT)
    except Exception as e:
        return False, f"request raised: {e}"

    if not text.strip() or ref_area not in text:
        return False, f"no rows for REF_AREA={ref_area} in response ({len(text)} bytes total)"

    return True, f"OK -- {len(text)} bytes returned, REF_AREA={ref_area} present"


def _discover() -> int:
    """Fetch the WS_EER dataflow structure and print its dimension list --
    the honest way to find out whether a weight-bearing dimension exists,
    rather than guessing the query-key syntax (see module docstring)."""
    try:
        import httpx
        resp = httpx.get(BIS_EER_DATAFLOW_STRUCTURE_URL, params={"references": "all"}, timeout=30.0)
        resp.raise_for_status()
    except Exception as e:
        print(f"FAIL: could not fetch dataflow structure: {e}")
        return 1

    body = resp.text
    print(f"Fetched {len(body)} bytes of dataflow structure from {BIS_EER_DATAFLOW_STRUCTURE_URL}")
    print("Inspect manually for a weight-bearing dimension/codelist (e.g. anything")
    print("resembling 'WEIGHT', 'BASKET_SHARE', or similar) alongside the standard")
    print("FREQ/TYPE(N,R)/BASKET(B,N)/REF_AREA dimensions -- absence of one there")
    print("is the concrete signal that weights are a documentation artifact, not")
    print("an API-queryable series (see module docstring).")
    return 0


def _discover_weights() -> int:
    """Download BIS's actual published Broad EER weights file and print its
    real internal structure -- Gate 1's answer, found this thread: weights
    are NOT in the SDMX API at all, they're a plain .xlsx download linked
    from data.bis.org/topics/EER's own Methodology section. This function
    does NOT assume a row/column layout (that would just be a new version
    of the same guessing problem this whole thread has been fixing) -- it
    downloads, opens the workbook, and reports sheet names/dimensions plus
    a structural sample and a scan for our own currency/REF_AREA codes, so
    a follow-up pass can write targeted extraction logic once the real
    layout is confirmed. Same two-phase discover-then-extract pattern as
    --discover for the API structure itself.
    """
    try:
        import httpx
        resp = httpx.get(BIS_EER_WEIGHTS_BROAD_URL, timeout=30.0, follow_redirects=True)
        resp.raise_for_status()
    except Exception as e:
        print(f"FAIL: could not download weights file: {e}")
        return 1

    print(f"Downloaded {len(resp.content)} bytes from {BIS_EER_WEIGHTS_BROAD_URL}")

    try:
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:
        print(f"FAIL: downloaded but could not parse as xlsx: {e}")
        return 1

    print(f"Sheets found: {wb.sheetnames}")

    # Search for our own currency/REF_AREA codes anywhere in each sheet's
    # first 200 rows -- bounded scan, since we don't know the layout and
    # some of these weight files span decades of vintages.
    target_codes = set(BROAD_DOLLAR_REF_AREAS.keys()) | set(BROAD_DOLLAR_REF_AREAS.values())
    any_matches = False

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet {sheet_name!r}: {ws.max_row} rows x {ws.max_column} cols ---")

        sample_rows = min(10, ws.max_row or 0)
        sample_cols = min(15, ws.max_column or 0)
        if sample_rows and sample_cols:
            print(f"First {sample_rows} rows x {sample_cols} cols:")
            for row in ws.iter_rows(min_row=1, max_row=sample_rows, max_col=sample_cols, values_only=True):
                print(f"  {row}")

        scan_rows = min(200, ws.max_row or 0)
        matches: list[tuple[int, int, object]] = []
        if scan_rows:
            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, max_row=scan_rows, values_only=True), start=1
            ):
                for col_idx, cell in enumerate(row, start=1):
                    if cell is not None and str(cell).strip() in target_codes:
                        matches.append((row_idx, col_idx, cell))
        if matches:
            any_matches = True
            print(f"Currency/REF_AREA code matches (row, col, value): {matches[:40]}")
        else:
            print(f"No direct currency-code matches in the first {scan_rows} rows "
                  f"-- this sheet may use full country names instead of codes, "
                  f"or the data may start further down.")

    wb.close()

    if not any_matches:
        print("\nNo sheet matched a currency/REF_AREA code directly in the scanned")
        print("range. The file may use country names or a different code convention")
        print("-- inspect the printed row samples above manually before writing any")
        print("targeted extraction logic.")

    return 0


def _extract_weights(sheet: str | None = None, us_ref_area: str = US_REF_AREA) -> int:
    """Gate 1's actual answer: download weightsb.xlsx, select a vintage
    sheet (defaults to the most recent -- max(sheetnames), which resolves
    to '2020_2022' given the 4 Aug 2026 discovery run's confirmed sheet
    set and the "YYYY_YYYY" naming already sorting correctly by year),
    and print each of the 13 BROAD_DOLLAR_REF_AREAS currencies' weight in
    the United States' own Broad EER basket -- the empirical replacement
    for Architecture v2.0 §7.2's hand-approximated 10-pair BIS_WEIGHTS
    dict.

    Has been authored and unit-tested against a synthetic workbook only
    (extract_us_weights_from_sheet()'s own tests) -- like
    --discover-weights before it, this has not yet been run against the
    real file from any sandbox on this project (no route to bis.org).
    """
    try:
        import httpx
        resp = httpx.get(BIS_EER_WEIGHTS_BROAD_URL, timeout=30.0, follow_redirects=True)
        resp.raise_for_status()
    except Exception as e:
        print(f"FAIL: could not download weights file: {e}")
        return 1

    print(f"Downloaded {len(resp.content)} bytes from {BIS_EER_WEIGHTS_BROAD_URL}")

    try:
        from io import BytesIO
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(resp.content), read_only=True, data_only=True)
    except Exception as e:
        print(f"FAIL: downloaded but could not parse as xlsx: {e}")
        return 1

    target_sheet = sheet or max(wb.sheetnames)
    if target_sheet not in wb.sheetnames:
        print(f"FAIL: sheet {target_sheet!r} not found. Available: {wb.sheetnames}")
        wb.close()
        return 1

    ws = wb[target_sheet]
    weights = extract_us_weights_from_sheet(ws, BROAD_DOLLAR_REF_AREAS, us_ref_area=us_ref_area)
    wb.close()

    if weights is None:
        print(f"FAIL: no row with column 2 == {us_ref_area!r} found in sheet {target_sheet!r} "
              f"within the first {_EXTRACT_WEIGHTS_MAX_SCAN_ROWS} rows -- cannot extract "
              f"Broad Dollar weights without locating the United States row. If BIS uses a "
              f"different REF_AREA code for the US in this file, pass it explicitly with "
              f"--us-ref-area.")
        return 1

    print(f"\nSheet: {target_sheet}  (vintage; BIS revises this table on a 3-year cycle)")
    print(f"{'Currency':10s} {'REF_AREA':9s} Weight in US Broad EER basket (%)")
    missing = []
    for currency, ref_area in sorted(BROAD_DOLLAR_REF_AREAS.items()):
        val = weights.get(currency)
        if val is None:
            missing.append(currency)
            print(f"  {currency:8s} {ref_area:9s} MISSING")
        else:
            print(f"  {currency:8s} {ref_area:9s} {val:.6f}")

    if missing:
        print(f"\n{len(missing)}/{len(BROAD_DOLLAR_REF_AREAS)} currencies MISSING a weight "
              f"in the US row: {missing}")
        return 1

    total = sum(v for v in weights.values() if v is not None)
    print(f"\nSum of these 13 target-currency weights: {total:.6f}")
    print("(Out of the US row's full basket, which spans ~64 economies -- this sum")
    print("being well under 100 is expected, not a sign of a parsing error.)")
    print(f"\nAll {len(BROAD_DOLLAR_REF_AREAS)} target-currency weights extracted successfully.")
    print("Next step (not done by this script): wire these into a config-driven")
    print("BIS_WEIGHTS equivalent for gold/cross_asset/broad_dollar.py, replacing the")
    print("hand-approximated 10-pair dict in Architecture v2.0 §7.2.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--discover", action="store_true",
        help="Fetch the WS_EER dataflow structure instead of querying index values",
    )
    parser.add_argument(
        "--discover-weights", action="store_true",
        help="Download and inspect BIS's published Broad EER weights xlsx (Gate 1)",
    )
    parser.add_argument(
        "--extract-weights", action="store_true",
        help="Download weightsb.xlsx and extract the 13 target currencies' weights "
             "in the US Broad EER basket (Gate 1 -- targeted extraction)",
    )
    parser.add_argument(
        "--sheet", default=None,
        help="Vintage sheet to extract from with --extract-weights (e.g. 2017_2019). "
             "Defaults to the most recent vintage found in the workbook.",
    )
    parser.add_argument(
        "--us-ref-area", default=US_REF_AREA,
        help=f"REF_AREA code for the United States row, used with --extract-weights "
             f"(default: {US_REF_AREA!r})",
    )
    parser.add_argument("--currency", default=None, help="Only check one currency (e.g. KRW)")
    args = parser.parse_args()

    if args.extract_weights:
        return _extract_weights(sheet=args.sheet, us_ref_area=args.us_ref_area)

    if args.discover_weights:
        return _discover_weights()

    if args.discover:
        return _discover()

    targets = list(BROAD_DOLLAR_REF_AREAS.items())
    if args.currency:
        targets = [(c, r) for c, r in targets if c == args.currency]
        if not targets:
            print(f"No mapping for currency {args.currency!r}. Known: {list(BROAD_DOLLAR_REF_AREAS)}")
            return 1

    failures = []
    for currency, ref_area in sorted(targets):
        ok, msg = _check_one(ref_area)
        status = "PASS" if ok else "FAIL"
        print(f"[{status}] {currency:5s} (REF_AREA={ref_area})  {msg}")
        if not ok:
            failures.append(currency)

    print()
    if failures:
        print(f"{len(failures)}/{len(targets)} currencies FAILED: {failures}")
        return 1

    print(f"All {len(targets)} currencies' EER index data reachable.")
    print("NOTE: this confirms INDEX availability only. Gate 1's actual weight")
    print("COMPONENTS live in a separate file, not this API -- run with")
    print("--discover-weights to inspect BIS's published Broad weights, or")
    print("--extract-weights to pull the actual 13 target-currency values.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
